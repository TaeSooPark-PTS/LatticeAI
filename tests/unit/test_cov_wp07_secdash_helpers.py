"""The pure helpers behind the admin security dashboard.

These functions decide what an operator is allowed to see: the per-user risk
roll-up, the label shown next to an account, and the four export encoders.
Every encoder is a place a live credential could escape the process, so the
tests assert on the produced rows/bytes rather than on "it returned something".
"""

from __future__ import annotations

import io
import sys
from types import ModuleType
from typing import Any, Dict, List

import pytest

from latticeai.api.security_dashboard import (
    _csv_dump,
    _excel_dump,
    _pdf_report,
    _summarize_user_risk,
    _user_label,
    soft_mask,
)

LEAKED = "sk-ABCDEFGHIJKLMNOPQRSTUVWXYZ012345"


# ── labels ────────────────────────────────────────────────────────────────
def test_a_row_with_no_email_is_labelled_rather_than_left_blank():
    assert _user_label({}, None) == "Unknown"
    assert _user_label({}, "") == "Unknown"


def test_a_label_prefers_nickname_then_name_then_the_address_itself():
    users = {
        "a@x.com": {"nickname": "Ann", "name": "Ann Kim"},
        "b@x.com": {"name": "Bob Lee"},
        "c@x.com": {},
    }
    assert _user_label(users, "a@x.com") == "Ann"
    assert _user_label(users, "b@x.com") == "Bob Lee"
    assert _user_label(users, "c@x.com") == "c@x.com"
    assert _user_label(users, "ghost@x.com") == "ghost@x.com"


def test_soft_mask_of_empty_text_is_empty_rather_than_a_row_of_stars():
    assert soft_mask("") == ""
    assert soft_mask("ab") == "**"


# ── per-user risk roll-up ─────────────────────────────────────────────────
def _classify(item: Dict[str, Any], index: int) -> Dict[str, Any]:
    content = str(item.get("content") or "")
    if content == "boom":
        raise RuntimeError("classifier is down")
    if "rrn" in content:
        return {"sensitivity": "high", "index": index}
    if "email" in content:
        return {"sensitivity": "low", "index": index}
    return {"sensitivity": "none", "index": index}


def _rollup() -> Dict[str, Dict[str, Any]]:
    history: List[Dict[str, Any]] = [
        {"role": "assistant", "content": "rrn is not counted for the model"},
        {
            "role": "user",
            "user_email": "a@x.com",
            "content": "rrn 900101-1234567",
            "timestamp": "2026-01-01T09:00:00",
        },
        {
            "role": "user",
            "user_email": "a@x.com",
            "content": "email me@x.com",
            "timestamp": "2026-01-02T09:00:00",
        },
        {
            "role": "user",
            "user_email": "a@x.com",
            "user_nickname": "Ann",
            "content": "just a question",
        },
        {"role": "user", "content": "boom"},
    ]
    file_events: List[Dict[str, Any]] = [
        {"user_email": "a@x.com", "sensitivity": "high", "filename": "payroll.xlsx"},
        {"user_email": "a@x.com", "sensitivity": "none", "sensitive_labels": ["secret"]},
        {"user_email": "b@x.com", "filename": "notes.txt"},
    ]
    rows = _summarize_user_risk(history, file_events, _classify)
    return {row["email"]: row for row in rows}


def test_only_user_turns_are_counted_and_the_worst_offender_sorts_first():
    rows = _summarize_user_risk(
        [
            {"role": "assistant", "content": "rrn 900101-1234567"},
            {"role": "user", "user_email": "a@x.com", "content": "rrn 900101-1234567"},
            {"role": "user", "user_email": "b@x.com", "content": "hello"},
        ],
        [],
        _classify,
    )
    assert [row["email"] for row in rows] == ["a@x.com", "b@x.com"]
    assert rows[0]["total_chats"] == 1, "the assistant turn must not be attributed to a user"


def test_risky_and_clean_chats_are_counted_separately_per_user():
    by_email = _rollup()
    ann = by_email["a@x.com"]
    assert ann["user"] == "Ann", "a turn that carries a nickname names the bucket"
    assert ann["total_chats"] == 3
    assert ann["risky_chats"] == 2
    assert ann["compliant_chats"] == 1
    assert ann["last_activity_at"] == "2026-01-02T09:00:00", "the newest timestamp wins"


def test_a_classifier_failure_is_counted_as_clean_rather_than_dropping_the_turn():
    by_email = _rollup()
    unknown = by_email["Unknown"]
    assert unknown["total_chats"] == 1
    assert unknown["risky_chats"] == 0
    assert unknown["compliant_chats"] == 1


def test_uploaded_files_are_risky_by_sensitivity_or_by_label():
    by_email = _rollup()
    ann = by_email["a@x.com"]
    assert ann["uploaded_files"] == 2
    assert ann["risky_files"] == 2, "a file flagged only by label is still risky"
    assert ann["compliant_files"] == 0
    assert by_email["b@x.com"]["compliant_files"] == 1


def test_high_severity_chats_and_files_drive_the_high_risk_counter():
    by_email = _rollup()
    assert by_email["a@x.com"]["high_risk_events"] == 2
    assert by_email["b@x.com"]["high_risk_events"] == 0


def test_the_risk_rate_is_a_percentage_of_everything_the_user_did():
    by_email = _rollup()
    # 2 risky chats + 2 risky files out of 3 chats + 2 files.
    assert by_email["a@x.com"]["risk_rate"] == 80.0
    assert by_email["b@x.com"]["risk_rate"] == 0.0


# ── CSV ───────────────────────────────────────────────────────────────────
def test_an_empty_csv_export_is_empty_bytes_rather_than_a_stray_header():
    assert _csv_dump([]) == b""


def test_csv_headers_are_the_union_of_the_rows_in_first_seen_order():
    out = _csv_dump([{"a": 1, "b": 2}, {"b": 3, "c": 4}]).decode("utf-8")
    assert out.splitlines()[0] == "a,b,c"
    assert out.splitlines()[2].startswith(",3,4")


def test_csv_values_are_redacted_even_when_the_key_looks_harmless():
    out = _csv_dump([{"note": "api_key=" + LEAKED, "count": 3}]).decode("utf-8")
    assert LEAKED not in out
    assert "[REDACTED_SECRET]" in out
    assert ",3" in out, "non-string values survive the export unchanged"


# ── Excel ─────────────────────────────────────────────────────────────────
def _read_sheet(blob: bytes) -> List[List[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob))
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


def test_an_excel_export_carries_the_header_union_and_masks_secrets():
    blob = _excel_dump([
        {"user": "alice", "note": "api_key=" + LEAKED},
        {"note": "clean", "count": 7},
    ])
    assert blob[:2] == b"PK", "an xlsx file is a zip container"
    rows = _read_sheet(blob)
    assert rows[0] == ["user", "note", "count"]
    assert LEAKED not in str(rows)
    assert "[REDACTED_SECRET]" in str(rows[1][1])
    assert rows[2][2] == 7


def test_an_empty_excel_export_says_so_instead_of_producing_no_sheet():
    assert _read_sheet(_excel_dump([])) == [["empty"]]


# ── PDF ───────────────────────────────────────────────────────────────────
class _Flowable:
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        self.args = args
        self.kwargs = kwargs

    @property
    def text(self) -> str:
        return str(self.args[0]) if self.args else ""


class _Paragraph(_Flowable):
    pass


class _Spacer(_Flowable):
    pass


class _Table(_Flowable):
    pass


@pytest.fixture()
def rendered(monkeypatch):
    """A recording stand-in for reportlab, so the render path runs everywhere.

    reportlab is not a declared dependency — ``pyproject.toml`` ships openpyxl
    but not reportlab — so on a clean install ``_pdf_report`` takes its
    library-missing fallback and the rendering branch never executes. Injecting
    the modules makes that branch reachable on every platform, and recording the
    flowables lets the test assert what the report was actually built from
    instead of "the bytes start with %PDF".
    """
    recorder: Dict[str, Any] = {}

    class _Doc:
        def __init__(self, buf: io.BytesIO, **kwargs: Any) -> None:
            self._buf = buf
            recorder["doc_kwargs"] = kwargs

        def build(self, story: List[Any]) -> None:
            recorder["story"] = list(story)
            self._buf.write(b"%PDF-1.4 recorded\n%%EOF\n")

    def _module(name: str, **attrs: Any) -> ModuleType:
        module = ModuleType(name)
        for key, value in attrs.items():
            setattr(module, key, value)
        return module

    fakes = {
        "reportlab": _module("reportlab"),
        "reportlab.lib": _module("reportlab.lib"),
        "reportlab.lib.pagesizes": _module("reportlab.lib.pagesizes", A4=(595.27, 841.89)),
        "reportlab.lib.styles": _module(
            "reportlab.lib.styles",
            getSampleStyleSheet=lambda: {
                "Title": "title-style",
                "Normal": "normal-style",
                "Heading2": "h2-style",
            },
        ),
        "reportlab.platypus": _module(
            "reportlab.platypus",
            Paragraph=_Paragraph,
            SimpleDocTemplate=_Doc,
            Spacer=_Spacer,
            Table=_Table,
        ),
    }
    for name, module in fakes.items():
        monkeypatch.setitem(sys.modules, name, module)
    return recorder


def test_a_pdf_report_titles_the_document_and_lists_the_overview(rendered):
    blob = _pdf_report(
        "Lattice AI Security Report",
        [{"event": "chat", "user": "a@x.com"}],
        {"total_messages": 2, "note": "api_key=" + LEAKED},
    )

    assert blob.startswith(b"%PDF-"), "the caller receives the renderer's bytes"
    assert rendered["doc_kwargs"]["title"] == "Lattice AI Security Report"
    texts = [f.text for f in rendered["story"] if isinstance(f, _Paragraph)]
    assert texts[0] == "<b>Lattice AI Security Report</b>"
    assert "total_messages: 2" in texts
    assert LEAKED not in " ".join(texts), "the overview is redacted before it is drawn"
    assert "note: api_key=[REDACTED_SECRET]" in texts


def test_a_pdf_report_tables_the_entries_and_caps_them_at_thirty(rendered):
    rows = [{"event": "chat", "note": "api_key=" + LEAKED} for _ in range(35)]
    _pdf_report("Security", rows, {})

    tables = [f for f in rendered["story"] if isinstance(f, _Table)]
    assert len(tables) == 1
    table_data = tables[0].args[0]
    assert table_data[0] == ["event", "note"], "the first row is the header"
    assert len(table_data) == 31, "30 entries plus the header"
    assert LEAKED not in str(table_data)
    assert table_data[1][1] == "api_key=[REDACTED_SECRET]"


def test_a_pdf_report_with_no_rows_draws_no_table(rendered):
    blob = _pdf_report("Empty", [], {"total_messages": 0})

    assert blob.startswith(b"%PDF-")
    assert not [f for f in rendered["story"] if isinstance(f, _Table)]
