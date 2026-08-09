"""wp16 coverage — ``lattice_brain.graph.discovery_index`` (local → graph).

This is the half of local ingestion that actually reads bytes: per-format text
extraction, the Computer/Drive/Folder hierarchy, the file-index upsert, graph
deletion, and the ``index_local_folder`` driver that decides — per file, per
run — indexed / skipped / too_large / failed / deleted.

Everything is driven against a real ``KnowledgeGraphStore`` over SQLite in
``tmp_path`` with real .pdf/.xlsx/.png inputs. Only unreadable-filesystem
behaviour is simulated, at the syscall seam.
"""

from __future__ import annotations

import json
import os
import sys
import types
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import lattice_brain.embeddings as embeddings_module
from lattice_brain.graph import _kg_fsutil as fsutil
from lattice_brain.graph.store import KnowledgeGraphStore


@pytest.fixture()
def store(tmp_path: Path) -> KnowledgeGraphStore:
    return KnowledgeGraphStore(tmp_path / "kg.sqlite", tmp_path / "blobs")


def _png(path: Path) -> Path:
    from PIL import Image

    Image.new("RGB", (12, 8), "blue").save(path)
    return path


def _pdf(path: Path) -> Path:
    from PIL import Image

    Image.new("RGB", (40, 20), "white").save(path, "PDF")
    return path


def _install_path_faults(monkeypatch, *, stat_fail=None, read_bytes_fail=None) -> None:
    stat_fail = stat_fail or {}
    read_bytes_fail = read_bytes_fail or {}
    real_stat = Path.stat
    real_is_dir = Path.is_dir
    real_is_file = Path.is_file
    real_read_bytes = Path.read_bytes

    def fake_stat(self, *args, **kwargs):
        exc = stat_fail.get(self.name)
        if exc is not None:
            raise exc
        return real_stat(self, *args, **kwargs)

    def fake_is_dir(self, *args, **kwargs):
        if self.name in stat_fail:
            return False
        return real_is_dir(self, *args, **kwargs)

    def fake_is_file(self, *args, **kwargs):
        if self.name in stat_fail:
            return True
        return real_is_file(self, *args, **kwargs)

    def fake_read_bytes(self, *args, **kwargs):
        exc = read_bytes_fail.get(self.name)
        if exc is not None:
            raise exc
        return real_read_bytes(self, *args, **kwargs)

    monkeypatch.setattr(Path, "stat", fake_stat)
    monkeypatch.setattr(Path, "is_dir", fake_is_dir)
    monkeypatch.setattr(Path, "is_file", fake_is_file)
    monkeypatch.setattr(Path, "read_bytes", fake_read_bytes)


def _index_row(store, source_id: str, relative_path: str):
    with store._connect() as conn:
        return conn.execute(
            "SELECT * FROM local_file_index WHERE source_id=? AND relative_path=?",
            (source_id, relative_path),
        ).fetchone()


def _register_source(store, source_id: str, root: Path) -> None:
    """local_file_index rows carry a FK to knowledge_sources."""
    with store._connect() as conn:
        conn.execute(
            """
            INSERT INTO knowledge_sources
              (id, root_path, os_type, drive_id, label, status, include_ocr,
               watch_enabled, consent_json, created_at, updated_at, last_scanned_at)
            VALUES (?, ?, 'macos', '/', 'root', 'active', 0, 0, '{}',
                    '2026-08-01T00:00:00Z', '2026-08-01T00:00:00Z', NULL)
            """,
            (source_id, str(root)),
        )


# ── _extract_local_file_text ─────────────────────────────────────────────────


def test_extract_text_from_pdf_records_page_count(store, tmp_path: Path) -> None:
    text, meta = store._extract_local_file_text(
        _pdf(tmp_path / "doc.pdf"), "pdf", include_ocr=False
    )

    assert meta["parser"] == "pdf"
    assert meta["pages"] == 1
    assert text == ""  # an image-only PDF has no extractable text


def test_extract_text_from_xlsx_skips_empty_rows_and_stops_at_the_budget(
    store, tmp_path: Path
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    first.title = "Q3"
    first["A1"] = "header"
    first["A3"] = "after a blank row"  # row 2 is entirely empty
    for row in range(4, 204):
        first.cell(row=row, column=1, value="x" * 1100)
    workbook.create_sheet("NeverRead")["A1"] = "unreachable"
    path = tmp_path / "book.xlsx"
    workbook.save(str(path))

    text, meta = store._extract_local_file_text(path, "spreadsheet", include_ocr=False)

    assert meta["sheets"] == 2
    assert meta["rows"] < 202  # the 200_000-character budget stopped the walk
    assert meta["cells"] == meta["rows"]
    assert text.startswith("[Sheet: Q3]")
    assert "after a blank row" in text
    assert "unreachable" not in text  # the second sheet is never reached


def test_extract_text_from_image_uses_the_vision_caption(
    store, tmp_path: Path
) -> None:
    path = _png(tmp_path / "pic.png")

    text, meta = store._extract_local_file_text(path, "image", include_ocr=False)

    assert meta["width"] == 12
    assert meta["height"] == 8
    assert meta["format"] == "PNG"
    assert meta["mode"] == "RGB"
    assert meta["ocr_enabled"] is False
    assert meta["vision_caption"] == "Image pic.png (PNG 12x8)"
    assert text == meta["vision_caption"]  # caption is the retrieval signal


def test_extract_text_from_image_runs_ocr_when_requested(
    monkeypatch, store, tmp_path: Path
) -> None:
    fake = types.ModuleType("pytesseract")
    fake.image_to_string = lambda image: "OCR: Lattice AI roadmap"
    monkeypatch.setitem(sys.modules, "pytesseract", fake)
    path = _png(tmp_path / "pic.png")

    text, meta = store._extract_local_file_text(path, "image", include_ocr=True)

    assert text == "OCR: Lattice AI roadmap"
    assert meta["ocr_enabled"] is True
    assert meta["ocr_chars"] == len("OCR: Lattice AI roadmap")
    assert meta["vision_caption"]  # caption still recorded alongside the OCR


def test_extract_text_from_image_survives_a_broken_vision_embedder(
    monkeypatch, store, tmp_path: Path
) -> None:
    def explode(*args, **kwargs):
        raise RuntimeError("vision backend unavailable")

    monkeypatch.setattr(embeddings_module, "get_vision_embedder", explode)
    path = _png(tmp_path / "pic.png")

    text, meta = store._extract_local_file_text(path, "image", include_ocr=False)

    assert text == ""
    assert meta["vision_caption"] == f"image:{path}"


# ── hierarchy / index-row helpers ────────────────────────────────────────────


def test_ensure_local_hierarchy_handles_files_outside_the_root(
    store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()

    with store._connect() as conn:
        parent_id = store._ensure_local_hierarchy(
            conn,
            source_id="src-1",
            root=root,
            file_path=tmp_path / "outside" / "note.md",
            os_type="macos",
            drive_id="/",
        )
        folders = {
            row["id"]: json.loads(row["metadata_json"])
            for row in conn.execute(
                "SELECT id, metadata_json FROM nodes WHERE type='Folder'"
            ).fetchall()
        }

    # No relative folders can be derived, so the file hangs off the root folder.
    assert list(folders) == [parent_id]
    assert folders[parent_id]["root"] is True
    assert folders[parent_id]["path"] == str(root)


def test_upsert_local_file_index_falls_back_to_the_file_name(
    store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    outside = tmp_path / "outside.md"
    outside.write_text("hello", encoding="utf-8")
    _register_source(store, "src-1", root)

    with store._connect() as conn:
        index_id = store._upsert_local_file_index(
            conn,
            source_id="src-1",
            root=root,
            file_path=outside,
            stat=outside.stat(),
            os_type="macos",
            drive_id="/",
            status="indexed",
            parser_type="plain_text",
        )
        row = conn.execute(
            "SELECT * FROM local_file_index WHERE id=?", (index_id,)
        ).fetchone()

    assert row["relative_path"] == "outside.md"
    assert row["size_bytes"] == 5
    assert row["last_indexed_at"]
    assert row["deleted"] == 0


def test_upsert_local_file_node_rejects_empty_text(store, tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()
    target = root / "note.md"
    target.write_text("hello", encoding="utf-8")

    with store._connect() as conn, pytest.raises(ValueError, match="비어 있습니다"):
        store._upsert_local_file_node(
            conn,
            source_id="src-1",
            root=root,
            file_path=target,
            stat=target.stat(),
            os_type="macos",
            drive_id="/",
            sha256="deadbeef",
            category="text",
            parser_type="plain_text",
            text="   ",
            parser_meta={},
        )


def test_upsert_local_file_node_handles_files_outside_the_root(
    store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    outside = tmp_path / "outside.md"
    outside.write_text("Lattice AI keeps notes.", encoding="utf-8")

    with store._connect() as conn:
        node_id = store._upsert_local_file_node(
            conn,
            source_id="src-1",
            root=root,
            file_path=outside,
            stat=outside.stat(),
            os_type="macos",
            drive_id="/",
            sha256="deadbeef",
            category="text",
            parser_type="plain_text",
            text="Lattice AI keeps notes.",
            parser_meta={"extracted_chars": 23},
        )
        metadata = json.loads(
            conn.execute(
                "SELECT metadata_json FROM nodes WHERE id=?", (node_id,)
            ).fetchone()["metadata_json"]
        )

    assert metadata["relative_path"] == "outside.md"
    assert metadata["file_path"] == str(outside)


# ── graph deletion / orphan cleanup ──────────────────────────────────────────


def test_delete_local_file_graph_ignores_a_missing_node_id(store) -> None:
    with store._connect() as conn:
        store._delete_local_file_graph(conn, None)
        assert conn.execute("SELECT COUNT(*) AS n FROM nodes").fetchone()["n"] == 0


def test_delete_local_file_graph_handles_a_childless_node(store) -> None:
    with store._connect() as conn:
        store._upsert_node(conn, "local-file:lonely", "Document", "lonely.md")
        store._delete_local_file_graph(conn, "local-file:lonely")
        assert conn.execute("SELECT COUNT(*) AS n FROM nodes").fetchone()["n"] == 0


def test_cleanup_local_graph_orphans_leaves_other_sources_alone(store) -> None:
    with store._connect() as conn:
        store._upsert_node(
            conn,
            "folder:other",
            "Folder",
            "other",
            metadata={"source_id": "other-source"},
        )
        store._cleanup_local_graph_orphans(conn, "mine")
        remaining = {
            row["id"] for row in conn.execute("SELECT id FROM nodes").fetchall()
        }

    assert remaining == {"folder:other"}


# ── small predicates ─────────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "parser,expected",
    [
        ("not-a-dict", False),
        ({"extracted_chars": "many"}, False),
        ({"extracted_chars": 0}, False),
        ({"extracted_chars": 12}, True),
    ],
)
def test_local_file_index_has_extracted_text(store, parser, expected) -> None:
    row = {"metadata_json": json.dumps({"parser": parser})}
    assert store._local_file_index_has_extracted_text(row) is expected


def test_node_matches_workspace_requires_a_node_id(store) -> None:
    with store._connect() as conn:
        assert store._node_matches_workspace(conn, None, None) is False
        assert store._node_matches_workspace(conn, "node:ghost", None) is False


# ── index_local_folder ───────────────────────────────────────────────────────


def test_index_local_folder_rejects_missing_paths(store, tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="경로가 존재하지 않습니다"):
        store.index_local_folder(tmp_path / "nope")


def test_index_local_folder_rejects_files(store, tmp_path: Path) -> None:
    target = tmp_path / "note.md"
    target.write_text("hi", encoding="utf-8")
    with pytest.raises(ValueError, match="폴더가 아닙니다"):
        store.index_local_folder(target)


def test_index_local_folder_reuses_the_existing_source_identity(
    store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    (root / "note.md").write_text("Lattice AI keeps notes.", encoding="utf-8")

    first = store.index_local_folder(root)
    second = store.index_local_folder(root, source_id_override="source:override")

    assert second["source"]["id"] == first["source"]["id"]
    with store._connect() as conn:
        ids = [
            row["id"] for row in conn.execute("SELECT id FROM knowledge_sources")
        ]
    assert ids == [first["source"]["id"]]


def test_index_local_folder_links_image_text_nodes(store, tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()
    _png(root / "pic.png")

    result = store.index_local_folder(root)

    assert result["counts"]["indexed"] == 1
    with store._connect() as conn:
        titles = {
            row["type"]: row["title"]
            for row in conn.execute(
                "SELECT type, title FROM nodes WHERE type IN ('Image', 'ImageText')"
            ).fetchall()
        }
    assert titles["Image"] == "pic.png"
    assert titles["ImageText"] == "pic.png OCR"


def test_index_local_folder_stops_at_max_files(store, tmp_path: Path) -> None:
    root = tmp_path / "root"
    root.mkdir()
    for index in range(3):
        (root / f"note{index}.md").write_text("Lattice AI notes.", encoding="utf-8")

    result = store.index_local_folder(root, max_files=1)

    assert result["counts"]["limit_reached"] == 1
    assert result["counts"]["indexed"] == 1
    assert "deleted" not in result["counts"]  # a partial scan deletes nothing


def test_index_local_folder_records_unreadable_files(
    monkeypatch, store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    (root / "locked.md").write_text("x", encoding="utf-8")
    _install_path_faults(
        monkeypatch, stat_fail={"locked.md": PermissionError("denied")}
    )

    result = store.index_local_folder(root)

    assert result["counts"]["failed"] == 1
    assert result["errors"][0]["path"].endswith("locked.md")
    assert result["errors"][0]["error"].startswith("permission_denied: ")


def test_index_local_folder_ignores_unknown_kinds_and_files_outside_the_root(
    monkeypatch, store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    outside = tmp_path / "outside.md"
    outside.write_text("Lattice AI keeps notes outside.", encoding="utf-8")
    # A file outside the root is classified by its *absolute* path, and macOS
    # temp directories live under /private — a sensitivity keyword. Emptying the
    # table keeps this test about the relative-path fallback rather than about
    # where the OS puts its temp files.
    monkeypatch.setattr(fsutil, "SENSITIVE_PATH_KEYWORDS", set())

    def fake_entries(scan_root, *, max_files):
        yield {"kind": "future_kind", "path": scan_root / "mystery.md"}
        yield {"kind": "file", "path": outside, "stat": outside.stat()}

    monkeypatch.setattr(store, "_iter_local_scan_entries", fake_entries)

    result = store.index_local_folder(root)

    assert result["counts"]["indexed"] == 1
    with store._connect() as conn:
        rows = conn.execute(
            "SELECT relative_path FROM local_file_index"
        ).fetchall()
    assert [row["relative_path"] for row in rows] == ["outside.md"]
    row = _index_row(store, result["source"]["id"], "outside.md")
    assert row["status"] == "indexed"
    assert row["file_path"] == str(outside)


def test_index_local_folder_drops_the_graph_when_a_file_becomes_unindexable(
    monkeypatch, store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    (root / "note.md").write_text("Lattice AI keeps notes.", encoding="utf-8")
    first = store.index_local_folder(root)
    source_id = first["source"]["id"]
    node_id = _index_row(store, source_id, "note.md")["graph_node_id"]
    assert node_id

    monkeypatch.setattr(fsutil, "LOCAL_SIZE_LIMITS", {"text": 1, "document": 1})
    second = store.index_local_folder(root)

    assert second["counts"]["too_large"] == 1
    row = _index_row(store, source_id, "note.md")
    assert row["status"] == "too_large"
    assert json.loads(row["metadata_json"])["reason"] == "size>1"
    with store._connect() as conn:
        assert (
            conn.execute(
                "SELECT COUNT(*) AS n FROM nodes WHERE id=?", (node_id,)
            ).fetchone()["n"]
            == 0
        )


def test_index_local_folder_records_a_read_failure(
    monkeypatch, store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    target = root / "note.md"
    target.write_text("Lattice AI keeps notes.", encoding="utf-8")
    first = store.index_local_folder(root)
    source_id = first["source"]["id"]
    node_id = _index_row(store, source_id, "note.md")["graph_node_id"]

    target.write_text("Lattice AI keeps different notes now.", encoding="utf-8")
    _install_path_faults(monkeypatch, read_bytes_fail={"note.md": OSError("disk error")})

    second = store.index_local_folder(root)

    assert second["counts"]["failed"] == 1
    assert second["errors"][0]["error"] == "disk error"
    row = _index_row(store, source_id, "note.md")
    assert row["status"] == "failed"
    assert row["error_message"] == "disk error"
    with store._connect() as conn:
        assert (
            conn.execute(
                "SELECT COUNT(*) AS n FROM nodes WHERE id=?", (node_id,)
            ).fetchone()["n"]
            == 0
        )


def test_index_local_folder_skips_a_touched_but_identical_file(
    store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    target = root / "note.md"
    target.write_text("Lattice AI keeps notes.", encoding="utf-8")
    first = store.index_local_folder(root)
    source_id = first["source"]["id"]
    node_id = _index_row(store, source_id, "note.md")["graph_node_id"]

    # Same bytes, new mtime: the cheap size/mtime check misses, the digest saves it.
    os.utime(target, (1_700_000_000, 1_700_000_000))
    second = store.index_local_folder(root)

    assert second["counts"]["skipped_unchanged"] == 1
    assert "indexed" not in second["counts"]
    row = _index_row(store, source_id, "note.md")
    assert row["graph_node_id"] == node_id
    assert json.loads(row["metadata_json"])["sha256_unchanged"] is True


def test_index_local_folder_records_an_extraction_failure(
    monkeypatch, store, tmp_path: Path
) -> None:
    root = tmp_path / "root"
    root.mkdir()
    target = root / "note.md"
    target.write_text("Lattice AI keeps notes.", encoding="utf-8")
    first = store.index_local_folder(root)
    source_id = first["source"]["id"]
    node_id = _index_row(store, source_id, "note.md")["graph_node_id"]

    target.write_text("Lattice AI keeps different notes now.", encoding="utf-8")

    def explode(*args, **kwargs):
        raise RuntimeError("parser crashed")

    monkeypatch.setattr(store, "_extract_local_file_text", explode)
    second = store.index_local_folder(root)

    assert second["counts"]["failed"] == 1
    assert second["errors"][0]["error"] == "parser crashed"
    row = _index_row(store, source_id, "note.md")
    assert row["status"] == "failed"
    assert row["error_message"] == "parser crashed"
    assert row["sha256"]  # the digest was known before extraction failed
    with store._connect() as conn:
        assert (
            conn.execute(
                "SELECT COUNT(*) AS n FROM nodes WHERE id=?", (node_id,)
            ).fetchone()["n"]
            == 0
        )
